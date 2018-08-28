from django.http import JsonResponse, HttpResponse
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import viewsets, permissions
from rest_framework.views import APIView

from api.views import CsrfExemptSessionAuthentication
from info.models import Student
from .models import Item, Experiment, Feedback, Grade
from .serializers import ItemSerializer, ExperimentSerializer, FeedbackSerializer, GradeSerializer


class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)


class ExperimentViewSet(viewsets.ModelViewSet):
    queryset = Experiment.objects.all()
    serializer_class = ExperimentSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Experiment.objects.all()
        item = self.request.query_params.get('item', '')
        course = self.request.query_params.get('course', '')
        if course:
            queryset = queryset.filter(course=course)
        if item:
            queryset = queryset.filter(item=item)
        return queryset


class FeedbackViewSet(viewsets.ModelViewSet):
    queryset = Feedback.objects.all()
    serializer_class = FeedbackSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Feedback.objects.all()
        item = self.request.query_params.get('item', None)
        if item:
            queryset = queryset.filter(experiment__item=item)
        return queryset


class GradeViewSet(viewsets.ModelViewSet):
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Grade.objects.all()
        experiment = self.request.query_params.get('experiment', None)
        if experiment:
            queryset = queryset.filter(experiment=experiment)
        return queryset


class GradeImportView(APIView):
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        user = request.user
        data = request.data.get('data')
        file = request.data.get('file')
        experiment = request.data.get('experiment')
        experiment = Experiment.objects.get(id=experiment)
        gradeList = Grade.objects.filter(experiment=experiment).all()
        xhList = [grade.student.xh for grade in gradeList]
        json = {'status': 'ok', 'data': [], 'warning': []}
        if file:
            workbook = load_workbook(file)
            worksheet = workbook[workbook.sheetnames[0]]
            for row in worksheet.rows:
                line = [col.value for col in row]
                if len(line) != 4:
                    json['status'] = 'error'
                    break
                if line[0] == '学号':
                    continue
                xh, name, grade, comment = line
                if grade:
                    if str(xh) in xhList:
                        json['warning'].append(name)
                    else:
                        json['data'].append({'xh': str(xh), 'name': name, 'grade': grade, 'comment': comment})
        if data:
            grades = []
            for d in data:
                student = Student.objects.get(xh=d['xh'])
                grade = Grade(experiment=experiment,
                              student=student,
                              grade=d['grade'],
                              comment=d['comment'])
                grades.append(grade)
            Grade.objects.bulk_create(grades)
        return JsonResponse(json)


class GradeImportTemplateView(APIView):
    authentication_classes = (CsrfExemptSessionAuthentication,)

    def get(self, request):
        classes = request.GET.get('classes')
        students = Student.objects.filter(classes=classes).all()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(['学号', '姓名', '成绩', '评语'])
        for student in students:
            ws.append([student.xh, student.name])
        response = HttpResponse(content=save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=template.xlsx'
        return response
