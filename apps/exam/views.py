import random
from datetime import datetime

from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import viewsets, permissions
from rest_framework.views import APIView

from api.views import CsrfExemptSessionAuthentication
from experiment.models import Experiment, Item
from .models import ExamSetting, Question, ExamRecord
from .serializers import ExamSettingSerializer, QuestionSerializer


class ExamSettingViewSet(viewsets.ModelViewSet):
    queryset = ExamSetting.objects.all()
    serializer_class = ExamSettingSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        user = self.request.user
        queryset = ExamSetting.objects.filter(teacher=user)
        date = self.request.query_params.get('date', None)
        if date:
            date = datetime.strptime(date, '%Y-%m-%d')
            queryset = queryset.filter(datetime__date=date)
        return queryset

    def create(self, request, *args, **kwargs):
        request.data['teacher'] = request.user.id
        experiment = Experiment.objects.get(id=request.data['experiment'])
        question_list = Question.objects.filter(item=experiment.item)
        if question_list.exists():
            if question_list.count() > 10:
                question_list = random.sample(list(question_list), 10)
            else:
                question_list = question_list.all()
            request.data['questions'] = ','.join([str(exam.id) for exam in question_list])
            return super().create(request, *args, **kwargs)
        else:
            return JsonResponse({'status': 'error', 'errMsg': '所选实验题库为空！'})


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Question.objects.all()
        item = self.request.query_params.get('item', None)
        if item:
            queryset = queryset.filter(item=item)
        return queryset


class QuestionImportView(APIView):
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        data = request.data.get('data')  # 确认后的数据
        file = request.data.get('file')  # 上传的文件
        item = request.data.get('item')
        item = Item.objects.get(id=item)
        json = {'status': 'ok', 'data': []}
        if file:
            workbook = load_workbook(file)
            worksheet = workbook[workbook.sheetnames[0]]
            for row in worksheet.rows:
                line = [col.value for col in row]
                if line[0] == '题目':
                    continue
                title, a, b, c, d, answer = line
                if not title:
                    continue
                json['data'].append({
                    'title': str(title),
                    'a': str(a),
                    'b': str(b),
                    'c': str(c),
                    'd': str(d),
                    'answer': str(answer),
                })
        if data:
            questions = list()
            for d in data:
                question = Question(title=d['title'], a=d['a'], b=d['b'], c=d['c'], d=d['d'], answer=d['answer'],
                                    item=item)
                questions.append(question)
            Question.objects.bulk_create(questions)
        return JsonResponse(json)


class Report(APIView):
    authentication_classes = (CsrfExemptSessionAuthentication,)

    def get(self, request, exam_id):
        exam = ExamSetting.objects.get(id=exam_id)
        record_list = ExamRecord.objects.filter(setting=exam)
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(['学号', '姓名', '得分', '提交时间'])
        for record in record_list:
            ws.append([record.student.xh, record.student.name, record.result, record.time.strftime('%H:%M:%S')])
        response = HttpResponse(content=save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=report.xlsx'
        return response
