from datetime import datetime

from django.contrib.auth import get_user_model, login, logout
from django.http import JsonResponse
from openpyxl import load_workbook
from rest_framework import viewsets, permissions, exceptions, status
from rest_framework.authentication import SessionAuthentication
from rest_framework.compat import authenticate
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView, exception_handler

from exam.models import ExamSetting, Question
from exam.serializers import ExamSettingSerializer, QuestionSerializer
from experiment.models import Experiment, Item, Feedback, Grade
from experiment.serializers import ExperimentSerializer, ItemSerializer, FeedbackSerializer, GradeSerializer
from info.models import User, Classes, Course, Student
from info.serializers import UserSerializer, ClassesSerializer, CourseSerializer, StudentSerializer
from thinking.models import Thinking
from thinking.serializers import ThinkingSerializer


def custom_exception_handler(exc, context):
    response = exception_handler(exc, context)

    if response is not None:
        if isinstance(context['view'], LoginView):
            # 登陆失败
            if isinstance(exc, exceptions.AuthenticationFailed):
                response.status_code = 200
                response.data['status'] = 'error'
                response.data['currentAuthority'] = 'guest'
                del response.data['detail']

    return response


class LoginView(APIView):
    def post(self, request):
        username = request.data.get('userName', None)
        password = request.data.get('password', None)

        if not username or not password:
            raise exceptions.AuthenticationFailed()

        credentials = {
            get_user_model().USERNAME_FIELD: username,
            'password': password
        }
        user = authenticate(**credentials)

        if user is None:
            raise exceptions.AuthenticationFailed()
        if not user.is_active:
            raise exceptions.AuthenticationFailed()

        login(request, user)
        return JsonResponse({
            'status': 'ok',
            'currentAuthority': 'admin' if user.is_admin else 'user'
        })


class LogoutView(APIView):
    def get(self, request):
        logout(request)
        return JsonResponse({
            'status': 'ok'
        })


class CurrentUserView(APIView):
    def get(self, request):
        user = request.user
        if isinstance(user, User):
            name = user.last_name + user.first_name

            return JsonResponse({
                'currentUser': {
                    'userid': user.id,
                    'name': name if name else user.username,
                    'last_name': user.last_name,
                    'first_name': user.first_name,
                },
                'currentAuthority': 'admin' if user.is_admin else 'user'
            })
        response = Response()
        response.status_code = status.HTTP_401_UNAUTHORIZED
        return response


# 暂时解决csrf问题
class CsrfExemptSessionAuthentication(SessionAuthentication):
    def enforce_csrf(self, request):
        return  # To not perform the csrf check previously happening


class SettingView(APIView):
    authentication_classes = (CsrfExemptSessionAuthentication,)

    def post(self, request):
        user = request.user
        password = request.data.get('password')
        password_new = request.data.get('password_new')
        last_name = request.data.get('last_name')
        first_name = request.data.get('first_name')
        json = {'status': 'ok'}
        if not user.check_password(password):
            json['status'] = 'error'
            return JsonResponse(json)
        user.last_name = last_name
        user.first_name = first_name
        if password_new:
            user.set_password(password_new)
        user.save()
        return JsonResponse(json)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    parser_classes = (FormParser, JSONParser, MultiPartParser)
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = User.objects.all()
        username = self.request.query_params.get('username', '')
        is_admin = self.request.query_params.get('is_admin', None)
        if is_admin is not None:  # str to bool
            is_admin = True if 'true' == is_admin else False
            queryset = queryset.filter(is_admin=is_admin)
        if username:
            queryset = queryset.filter(username__icontains=username)
        return queryset

    def create(self, request, *args, **kwargs):
        # 保存时生成用户名
        response = super().create(request, *args, **kwargs)
        if status.HTTP_201_CREATED == response.status_code:
            user = User.objects.get(username=request.data['username'])
            user.set_password('123456')
            user.save()
        return response

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)


class ClassesViewSet(viewsets.ModelViewSet):
    queryset = Classes.objects.all()
    serializer_class = ClassesSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Course.objects.all()
        name = self.request.query_params.get('name', '')
        _status = self.request.query_params.get('status', None)
        classes = self.request.query_params.get('classes', '')
        if classes:
            queryset = queryset.filter(classes=classes)
        if _status is not None:
            _status = True if 'true' == _status else False
            queryset = queryset.filter(status=_status)
        if name:
            queryset = queryset.filter(name__icontains=name)
        return queryset


class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)


class ExperimentViewSet(viewsets.ModelViewSet):
    queryset = Experiment.objects.all()
    serializer_class = ExperimentSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Experiment.objects.all()
        item = self.request.query_params.get('item', '')
        course = self.request.query_params.get('course', '')
        if course:
            queryset = queryset.filter(course=course)
        if item:
            queryset = queryset.filter(item=item)
        return queryset


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Student.objects.all()
        name = self.request.query_params.get('name', '')
        classes = self.request.query_params.get('classes', '')
        if classes:
            queryset = queryset.filter(classes=classes)
        if name:
            queryset = queryset.filter(name__icontains=name)
        return queryset


class ExamSettingViewSet(viewsets.ModelViewSet):
    queryset = ExamSetting.objects.all()
    serializer_class = ExamSettingSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        user = self.request.user
        queryset = ExamSetting.objects.filter(teacher=user)
        date = self.request.query_params.get('date', None)
        if date:
            date = datetime.strptime(date, '%Y-%m-%d')
            queryset = queryset.filter(datetime__date=date)
        return queryset

    def create(self, request, *args, **kwargs):
        request.data['teacher'] = request.user.id
        return super().create(request, *args, **kwargs)


class ThinkingViewSet(viewsets.ModelViewSet):
    queryset = Thinking.objects.all()
    serializer_class = ThinkingSerializer
    parser_classes = (MultiPartParser,)
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Thinking.objects.all()
        item = self.request.query_params.get('item', None)
        if item:
            queryset = queryset.filter(item=item)
        return queryset


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Question.objects.all()
        item = self.request.query_params.get('item', None)
        if item:
            queryset = queryset.filter(item=item)
        return queryset


class FeedbackViewSet(viewsets.ModelViewSet):
    queryset = Feedback.objects.all()
    serializer_class = FeedbackSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Feedback.objects.all()
        item = self.request.query_params.get('item', None)
        if item:
            queryset = queryset.filter(experiment__item=item)
        return queryset


class GradeViewSet(viewsets.ModelViewSet):
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Grade.objects.all()
        experiment = self.request.query_params.get('experiment', None)
        if experiment:
            queryset = queryset.filter(experiment=experiment)
        return queryset


class GradeImportView(APIView):
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        user = request.user
        data = request.data.get('data')
        file = request.data.get('file')
        experiment = request.data.get('experiment')
        experiment = Experiment.objects.get(id=experiment)
        gradeList = Grade.objects.filter(experiment=experiment).all()
        xhList = [grade.student.xh for grade in gradeList]
        json = {'status': 'ok', 'data': [], 'warning': []}
        if file:
            workbook = load_workbook(file)
            worksheet = workbook[workbook.sheetnames[0]]
            for row in worksheet.rows:
                line = [col.value for col in row]
                if len(line) != 4:
                    json['status'] = 'error'
                    break
                if line[0] == '学号':
                    continue
                xh, name, grade, comment = line
                if str(xh) in xhList:
                    json['warning'].append(name)
                else:
                    json['data'].append({'xh': str(xh), 'name': name, 'grade': grade, 'comment': comment})
        if data:
            grades = []
            for d in data:
                student = Student.objects.get(xh=d['xh'])
                grade = Grade(experiment=experiment,
                              student=student,
                              grade=d['grade'],
                              comment=d['comment'])
                grades.append(grade)
            Grade.objects.bulk_create(grades)
        return JsonResponse(json)
