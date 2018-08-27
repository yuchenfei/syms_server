from django.http import JsonResponse
from openpyxl import load_workbook
from rest_framework import viewsets, permissions, status
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.views import APIView

from api.views import CsrfExemptSessionAuthentication
from .models import User, Classes, Course, Student
from .serializers import UserSerializer, ClassesSerializer, CourseSerializer, StudentSerializer


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    parser_classes = (FormParser, JSONParser, MultiPartParser)
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = User.objects.all()
        username = self.request.query_params.get('username', '')
        is_admin = self.request.query_params.get('is_admin', None)
        if is_admin is not None:  # str to bool
            is_admin = True if 'true' == is_admin else False
            queryset = queryset.filter(is_admin=is_admin)
        if username:
            queryset = queryset.filter(username__icontains=username)
        return queryset

    def create(self, request, *args, **kwargs):
        # 保存时生成用户名
        response = super().create(request, *args, **kwargs)
        if status.HTTP_201_CREATED == response.status_code:
            user = User.objects.get(username=request.data['username'])
            user.set_password('123456')
            user.save()
        return response

    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)


class ClassesViewSet(viewsets.ModelViewSet):
    queryset = Classes.objects.all()
    serializer_class = ClassesSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Course.objects.all()
        name = self.request.query_params.get('name', '')
        _status = self.request.query_params.get('status', None)
        classes = self.request.query_params.get('classes', '')
        if classes:
            queryset = queryset.filter(classes=classes)
        if _status is not None:
            _status = True if 'true' == _status else False
            queryset = queryset.filter(status=_status)
        if name:
            queryset = queryset.filter(name__icontains=name)
        return queryset

    @staticmethod
    def handle_err_response(exc, context, response):
        if isinstance(exc, ValidationError):
            if exc.get_codes().get('non_field_errors')[0] == 'unique':
                response.status_code = 200
                response.data['errMsg'] = '创建的课程已存在'
        return response


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def get_queryset(self):
        queryset = Student.objects.all()
        name = self.request.query_params.get('name', '')
        classes = self.request.query_params.get('classes', '')
        if classes:
            queryset = queryset.filter(classes=classes)
        if name:
            queryset = queryset.filter(name__icontains=name)
        return queryset


class StudentImportView(APIView):
    authentication_classes = (CsrfExemptSessionAuthentication,)
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        user = request.user
        data = request.data.get('data')  # 确认后的数据
        file = request.data.get('file')  # 上传的文件
        classes = request.data.get('classes')
        classes = Classes.objects.get(id=classes)
        student_list = Student.objects.filter(classes=classes).all()
        xh_list = [student.xh for student in student_list]
        json = {'status': 'ok', 'data': [], 'warning': []}
        if file:
            workbook = load_workbook(file)
            worksheet = workbook[workbook.sheetnames[0]]
            for row in worksheet.rows:
                line = [col.value for col in row]
                if len(line) != 2:
                    json['status'] = 'error'
                    break
                if line[0] == '学号':
                    continue
                xh, name = line
                if str(xh) in xh_list:
                    json['warning'].append('{}({})'.format(name, xh))
                else:
                    json['data'].append({'xh': str(xh), 'name': name})
        if data:
            students = []
            for d in data:
                student = Student(xh=d['xh'], name=d['name'], classes=classes)
                students.append(student)
            Student.objects.bulk_create(students)
        print(json)
        return JsonResponse(json)
